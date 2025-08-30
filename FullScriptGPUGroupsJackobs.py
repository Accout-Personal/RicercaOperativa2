import json
import os
import copy
#import networkx as nx
import csv
import numpy as np

from numba import cuda
from pathlib import Path
from NFPLibGPU import *
from typing import Tuple
import cupy as cp
import gc
import networkit as nk
import pandas as pd
import time


import openpyxl

def update_or_add_with_openpyxl(filename, name, value):
    """Minimal update or add row using openpyxl with 2 decimal formatting"""
    
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    # Look for existing name in column 1
    found_row = None
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value and str(cell_value).lower() == str(name).lower():
            found_row = row
            break
    
    if found_row:
        # Update existing value in column 2
        target_row = found_row
        print(f"Updated {name}: {value}")
    else:
        # Add new row
        target_row = ws.max_row + 1
        ws.cell(row=target_row, column=3, value=name)
        print(f"Added {name}: {value}")
    
    # Write value
    value_cell = ws.cell(row=target_row, column=3, value=value)
    if isinstance(value, float):
        value_cell.number_format = '0.00'
    
    wb.save(filename)

MAX_NFP_VERTICES = 30

def remove_symmetric_duplicates_gpu(arr_cpu):
    """
    GPU-accelerated removal of symmetric duplicates
    
    Parameters:
        arr_cpu: numpy array of shape (N,2)
    Returns:
        numpy array with one of each symmetric pair removed
    """
    # Transfer data to GPU
    arr = cp.asarray(arr_cpu)
    
    # Create a canonical form where first element is always smaller
    mask = arr[:, 0] <= arr[:, 1]
    canonical = cp.zeros_like(arr)
    
    # Where mask is True, keep original order
    canonical[mask] = arr[mask]
    # Where mask is False, reverse the pairs
    canonical[~mask] = arr[~mask][:, ::-1]
    
    # Create unique identifiers using canonical form
    max_val = int(cp.max(arr)) + 1
    pair_ids = canonical[:, 0] * max_val + canonical[:, 1]
    
    # Find first occurrence of each unique pair
    _, indices = cp.unique(pair_ids, return_index=True)
    
    # Get original pairs at these indices and sort indices for consistency
    indices = cp.sort(indices)
    result_gpu = arr[indices]
    
    # Transfer back to CPU
    return cp.asnumpy(result_gpu)

def remove_symmetric_duplicates_cpu(arr):
    """
    CPU version for comparison, using same logic as GPU version
    """
    # Create canonical form where first element is always smaller
    mask = arr[:, 0] <= arr[:, 1]
    canonical = np.zeros_like(arr)
    
    # Where mask is True, keep original order
    canonical[mask] = arr[mask]
    # Where mask is False, reverse the pairs
    canonical[~mask] = arr[~mask][:, ::-1]
    
    # Create unique identifiers using canonical form
    max_val = int(np.max(arr)) + 1
    pair_ids = canonical[:, 0] * max_val + canonical[:, 1]
    
    # Find first occurrence of each unique pair
    _, indices = np.unique(pair_ids, return_index=True)
    
    # Sort indices for consistency
    indices.sort()
    return arr[indices]

@cuda.jit(device=True)
def on_segment(Ax,Ay, Bx,By, p):
    #p[0] => p['x'], p[1]=>p['y']
    # vertical line
    if almost_equal(Ax, Bx) and almost_equal(p[0], Ax):
        if not almost_equal(p[1], By) and not almost_equal(p[1], Ay) and p[1] < max(By, Ay) and p[1] > min(By, Ay):
            return True
        else:
            return False

    # horizontal line
    if almost_equal(Ay, By) and almost_equal(p[1], Ay):
        if not almost_equal(p[0], Bx) and not almost_equal(p[0], Ax) and p[0] < max(Bx, Ax) and p[0] > min(Bx, Ax):
            return True
        else:
            return False

    # range check
    if (p[0] < Ax and p[0] < Bx) or (p[0] > Ax and p[0] > Bx) or (p[1] < Ay and p[1] < By) or (p[1] > Ay and p[1] > By):
        return False

    # exclude end points
    if (almost_equal(p[0], Ax) and almost_equal(p[1], Ay)) or (almost_equal(p[0], Bx) and almost_equal(p[1], By)):
        return False

    cross = (p[1] - Ay) * (Bx - Ax) - (p[0] - Ax) * (By - Ay)

    if abs(cross) > 1e-9:
        return False

    dot = (p[0] - Ax) * (Bx - Ax) + (p[1] - Ay) * (By - Ay)

    if dot < 0 or almost_equal(dot, 0):
        return False

    len2 = (Bx - Ax)**2 + (By - Ay)**2

    if dot > len2 or almost_equal(dot, len2):
        return False

    return True

@cuda.jit(device=True)
def almost_equal(a, b):
    return abs(a - b) < 1e-9


@cuda.jit(device=True)
def point_in_polygon(point, polygon,number_vertices):
    
    # return True if point is in the polygon, False if outside, and None if exactly on a point or edge
    #False =>0, True =>1, None=>2
    inside = 0

    offsetx = 0
    offsety = 0
    #0 => x 1=>y
    j = number_vertices - 1
    for i in range(number_vertices):
        xi = polygon[i][0] + offsetx
        yi = polygon[i][1] + offsety
        xj = polygon[j][0] + offsetx
        yj = polygon[j][1] + offsety

        if almost_equal(xi, point[0]) and almost_equal(yi, point[1]):
            return 2  # no result
        
        if on_segment(xi, yi, xj, yj, point):
            return 2  # exactly on the segment
        
        if almost_equal(xi, xj) and almost_equal(yi, yj):  # ignore very small lines
            continue
        
        intersect = ((yi > point[1]) != (yj > point[1])) and (point[0] < (xj - xi) * (point[1] - yi) / (yj - yi) + xi)
        if intersect:
            #inside = not inside
            if inside==0:
                inside=1
            else:
                inside=0
        
        j = i
    #print("point in polygon check")
    #print(polygon)
    #print(point)
    #print(inside)
    return inside

@cuda.jit
def compare_groups_kernel(all_points_a, all_points_b, 
                         all_nfps, nfp_sizes, nfp_offsets,
                         batch_references,      # Add this - references for each group A in batch
                         pair_offsets_a, pair_offsets_b,
                         pair_sizes_a, pair_sizes_b,
                         num_pairs_in_batch,
                         results,pair_result_offsets,max_results):
    """
    Process batch of pairs
    Memory optimized version using uint16 for counts and offsets
    NFPs handled with offsets and shifted by references
    """
    #One block handles 1 point to one group of points
    #Pair ID
    pair_idx = cuda.blockIdx.y

    #print(f"Block ({cuda.blockIdx.x}, {cuda.blockIdx.y}) starting")  # Add this BEFORE first check

    if pair_idx >= num_pairs_in_batch:
    #    print(f"Block ({cuda.blockIdx.x}, {cuda.blockIdx.y}) rejected: pair_idx {pair_idx} >= {num_pairs_in_batch}")
        return
    
    #Block ID
    point_idx = cuda.blockIdx.x
    if point_idx >= pair_sizes_a[pair_idx]:
    #    print(f"Block ({cuda.blockIdx.x}, {cuda.blockIdx.y}) rejected: point_idx {point_idx} >= {pair_sizes_a[pair_idx]}")
        return
    
        # Debug print for block identification
    #if cuda.threadIdx.x == 0:
    #    print(f"Executing block ({cuda.blockIdx.x}, {cuda.blockIdx.y})")
    #    print(f"Points A size for this pair: {pair_sizes_a[pair_idx]}")
    #    print(f"Points B size for this pair: {pair_sizes_b[pair_idx]}")
    #    print(f"Offsets - A: {pair_offsets_a[pair_idx]}, B: {pair_offsets_b[pair_idx]}")

    # Get reference for this pair (based on group A)
    ref_x = batch_references[pair_idx, 0]
    ref_y = batch_references[pair_idx, 1]
    
    # Get our point from A
    point_a = all_points_a[pair_offsets_a[pair_idx] + point_idx]

    # Load NFP for this pair into shared memory using offsets and shift by reference
    shared_nfp = cuda.shared.array(shape=(MAX_NFP_VERTICES,2), dtype=np.float32)
    if cuda.threadIdx.x == 0:
        nfp_size = nfp_sizes[pair_idx]
        nfp_start = nfp_offsets[pair_idx]
        for i in range(nfp_size):
            # Shift NFP by reference and point A
            shared_nfp[i,0] = all_nfps[nfp_start + i,0] - ref_x + point_a[0]
            shared_nfp[i,1] = all_nfps[nfp_start + i,1] - ref_y + point_a[1]
    
    cuda.syncthreads()
    
    b_offset = pair_offsets_b[pair_idx]
    points_b_size = pair_sizes_b[pair_idx]
    # Get base offset for this pair's results
    pair_base = pair_result_offsets[pair_idx]

    # Each thread takes a continuous chunk of points_b
    thread_idx = cuda.threadIdx.x
    num_threads = cuda.blockDim.x
    points_per_thread = (points_b_size + num_threads - 1) // num_threads

    # Calculate start and end for this thread's chunk
    start_j = thread_idx * points_per_thread
    end_j = min(start_j + points_per_thread, points_b_size)

    for j in range(start_j,end_j):
        point_b = all_points_b[b_offset + j]
        result_idx = pair_base + (point_idx * points_b_size + j)
        
        if result_idx < max_results:
            #print(result_idx)
            results[result_idx, 0] = point_a[2]
            results[result_idx, 1] = point_b[2]
            results[result_idx, 2] = point_in_polygon(point_b, shared_nfp, nfp_sizes[pair_idx])


    
def process_group_pairs(all_pairs,groups,group_poly_ids,nfps,nfp_sizes, references,batch_size=200):

    """
    Process multiple pairs in batches
    all_pairs: list of (group_a_idx, group_b_idx) to process
    groups: list of point arrays for each group
    nfps: dictionary or array of NFPs for each pair
    """

    # Prepare results dictionary
    results_dict = {}
    
    # Process in batches
    for batch_start in range(0, len(all_pairs), batch_size):
        #cuda.current_context().synchronize()
        batch_end = min(batch_start + batch_size, len(all_pairs))
        batch_pairs = all_pairs[batch_start:batch_end]
        num_pairs = len(batch_pairs)
        # Prepare concatenated arrays for this batch

        # Calculate result offsets for each pair
        pair_result_offsets = np.zeros(num_pairs, dtype=np.int64)
        total_results = 0
        for i, (group_a_idx, group_b_idx) in enumerate(batch_pairs):
            pair_result_offsets[i] = total_results
            total_results += len(groups[group_a_idx]['InnerFitPoints']) * len(groups[group_b_idx]['InnerFitPoints'])
        #print("group a: ", groups[group_a_idx]['InnerFitPoints'], "group b: ", groups[group_b_idx]['InnerFitPoints'])
        #print("len group a: ", len(groups[group_a_idx]['InnerFitPoints']), "len group b: ", len(groups[group_b_idx]['InnerFitPoints']))
        #print("total results: ",total_results)
        points_a = []
        points_b = []
        batch_nfps = []
        batch_nfp_sizes = []
        pair_offsets_a = np.zeros(num_pairs, dtype=np.uint32) # 0-65535 range
        pair_offsets_b = np.zeros(num_pairs, dtype=np.uint32)
        nfp_offsets = np.zeros(num_pairs, dtype=np.uint32)
        pair_sizes_a = np.zeros(num_pairs, dtype=np.uint32)
        pair_sizes_b = np.zeros(num_pairs, dtype=np.uint32)
        batch_references = np.zeros((num_pairs, 2), dtype=np.float32)  # 2D for x,y coordinates
        # Calculate offsets and concatenate data
        offset_a = 0
        offset_b = 0
        nfp_offset = 0
        print(f"Processing batch {batch_start//batch_size + 1}, pairs {batch_start} to {batch_end}")
        for i, (group_a_idx, group_b_idx) in enumerate(batch_pairs):
            #Get Referecens 
            batch_references[i] = references[group_poly_ids[group_a_idx]]
            # Get points for this pair
            group_a_points = groups[group_a_idx]['InnerFitPoints']
            group_b_points = groups[group_b_idx]['InnerFitPoints']
            
            # Store offsets and sizes
            pair_offsets_a[i] = offset_a
            pair_offsets_b[i] = offset_b
            pair_sizes_a[i] = len(group_a_points)
            pair_sizes_b[i] = len(group_b_points)
            
            # Concatenate points
            points_a.append(group_a_points)
            points_b.append(group_b_points)
            
            # Get NFP for this pair
            nfp = nfps[(group_poly_ids[group_a_idx], group_poly_ids[group_b_idx])]
            nfp_size = nfp_sizes[(group_poly_ids[group_a_idx], group_poly_ids[group_b_idx])]
            nfp_offsets[i] = nfp_offset
            nfp_offset += len(nfp)
            batch_nfps.append(nfp)
            batch_nfp_sizes.append(nfp_size)

            # Update offsets
            offset_a += len(group_a_points)
            offset_b += len(group_b_points)
            
        #print("nfp: ",nfp)
        #print("nfp sizes ",nfp_size)
        #print("references A",references)
        # Convert lists to arrays
        all_points_a = np.concatenate(points_a).astype(np.float32)
        all_points_b = np.concatenate(points_b).astype(np.float32)
        all_nfps = np.concatenate(batch_nfps).astype(np.float32)
        

        batch_results = np.zeros((total_results, 3), dtype=np.uint32)
        
        # Transfer all data to GPU at once
        d_points_a = cuda.to_device(all_points_a)
        d_points_b = cuda.to_device(all_points_b)
        d_nfps = cuda.to_device(all_nfps)
        d_nfp_sizes = cuda.to_device(batch_nfp_sizes)
        d_nfp_offsets = cuda.to_device(nfp_offsets)
        d_references = cuda.to_device(batch_references)
        d_offsets_a = cuda.to_device(pair_offsets_a)
        d_offsets_b = cuda.to_device(pair_offsets_b)
        d_sizes_a = cuda.to_device(pair_sizes_a)
        d_sizes_b = cuda.to_device(pair_sizes_b)
        d_results = cuda.to_device(batch_results)
        d_pair_result_offsets = cuda.to_device(pair_result_offsets)
    
        # Configure kernel
        threadsperblock = 256
        max_points = int(max(pair_sizes_a))  # Convert uint16 to int for grid calculation
        #blockspergrid = (max_points, num_pairs)
        blockspergrid_x = max_points  # blocks for points_a
        blockspergrid_y = batch_size  # current batch size
        #print(f"Launching grid: ({blockspergrid_x}, {blockspergrid_y})")

        #print("GPU starts")
        # Launch kernel for whole batch
        compare_groups_kernel[(blockspergrid_x, blockspergrid_y), threadsperblock](
            d_points_a, d_points_b,
            d_nfps, d_nfp_sizes,d_nfp_offsets,
            d_references,
            d_offsets_a, d_offsets_b,
            d_sizes_a, d_sizes_b,
            num_pairs,
            d_results,d_pair_result_offsets,total_results
        )
        # Get results
        batch_results = d_results.copy_to_host()

        #print("GPU ends")
        ## Split results back into pairs
        #print("filtering results..")
        # Split results using the offsets
        for i, (group_a_idx, group_b_idx) in enumerate(batch_pairs):
            size = pair_sizes_a[i] * pair_sizes_b[i]
            start_idx = pair_result_offsets[i]
            end_idx = start_idx + size
            results_dict[(group_a_idx, group_b_idx)] = batch_results[start_idx:end_idx]
        for (group_a_idx, group_b_idx) in (batch_pairs):
            pair_results = results_dict[(group_a_idx, group_b_idx)]
            pair_results = pair_results[pair_results[:,2]==1][:,:2]
            results_dict[(group_a_idx, group_b_idx)] = pair_results


    return results_dict

if __name__ == "__main__":
    
    dir_path = os.path.dirname(os.path.realpath(__file__))

    jakobs1 = "/datasetJson/json/jakobs1.json"
    jakobs2 = "/datasetJson/json/jakobs2.json"

    #dataset struct:
    #{
    #   "name_polygon":{
    #       "VERTICES":[list_of_vertices],
    #       "QUANTITY":quantity,
    #       "NUMBER OF VERTICES": number
    #
    #}
    ##J

    Jackobs1 = {
        "J1-10-10-0":{"set":[1,5, 9, 12,15,16,17,20,21,23],"board":(10,20)},
        "J1-10-10-1":{"set":[1,3, 4, 5, 8, 9, 10,13,14,25],"board":(10,19)},
        "J1-10-10-2":{"set":[9,12,13,15,16,17,20,5, 2, 21],"board":(10,21)},
        "J1-10-10-3":{"set":[9,14,20,25,17,24,8, 22,6, 4],"board":(10,23)},
        "J1-10-10-4":{"set":[1,2, 3, 16,17,9, 22,10,8, 25],"board":(10,15)},

        "J1-12-20-0":{"set":[9,12,16,17,1,22,8,21,15,13,23,5],"board":(20,12)},
        "J1-12-20-1":{"set":[9,12,16,17,1,3,4,14,15,10,5,25],"board":(20,11)},
        "J1-12-20-2":{"set":[9,16,17,13,14,20,22,21,24,2,4,8],"board":(20,14)},
        "J1-12-20-3":{"set":[1,2,3,16,17,25,6,20,22,10,8,7],"board":(20,10)},
        "J1-12-20-4":{"set":[24,6,16,3,9,7,15,17,23,13,21,1],"board":(20,16)},

        "J1-14-20-0":{"set":[9,17,22,23,5,3,21,1,16,10,8,15,12,13],"board":(20,14)},
        "J1-14-20-1":{"set":[14,20,4,2,3,9,15,10,25,5,12,17,16,1],"board":(20,14)},
        "J1-14-20-2":{"set":[2,24,13,6,21,14,8,25,20,16,9,4,17,10],"board":(20,16)},
        "J1-14-20-3":{"set":[15,13,1,2,17,23,3,8,19,22,16,21,7,10],"board":(20,12)},
        "J1-14-20-4":{"set":[24,6,25,3,11,15,4,9,10,16,13,12,17,21],"board":(20,16)},

    }

    Jackobs2 = {
        "J2-10-35-0":{"set":[23,20,19,1,12,5,21,9,15,18],"board":(35,28)},
        "J2-10-35-1":{"set":[25,13,8,10,4,1,3,12,5,14],"board":(35,28)},
        "J2-10-35-2":{"set":[13,17,12,20,16,5,19,15,9,2],"board":(35,27)},
        "J2-10-35-3":{"set":[25,20,8,24,22,21,12,4,6,10],"board":(35,25)},
        "J2-10-35-4":{"set":[25,20,1,16,8,17,7,2,10,3],"board":(35,22)},

        "J2-12-35-0":{"set":[13,1,9,12,15,5,19,21,20,23,8,18],"board":(35,31)},
        "J2-12-35-1":{"set":[1,25,16,12,5,10,9,17,3,15,4,14],"board":(35,29)},
        "J2-12-35-2":{"set":[12,13,10,20,16,4,19,24,2,8,21,22],"board":(35,30)},
        "J2-12-35-3":{"set":[2,20,7,6,1,3,16,22,10,8,17,25],"board":(35,25)},
        "J2-12-35-4":{"set":[24,23,6,13,12,15,7,18,1,19,21,3],"board":(35,29)},

        "J2-14-35-0":{"set":[13,18,15,12,8,21,10,9,1,20,23,5,3,19],"board":(35,34)},
        "J2-14-35-1":{"set":[10,5,1,4,16,17,3,15,14,9,20,12,25,2],"board":(35,33)},
        "J2-14-35-2":{"set":[10,13,16,4,14,12,20,25,19,22,8,21,24,6],"board":(35,33)},
        "J2-14-35-3":{"set":[18,15,10,2,19,1,8,3,17,13,23,20,7,21],"board":(35,29)},
        "J2-14-35-4":{"set":[13,6,3,16,19,11,12,15,9,21,10,25,24,4],"board":(35,31)},
    }


    #Dataset selected for graph generation
    selelected = jakobs1
    selectSet = Jackobs1
    filepath = dir_path+selelected
    with open(filepath) as file:
            inputdataset = json.load(file)

    #Name of output
    metaResults = []
    for key,value in selectSet.items():

        proc_start = time.perf_counter()
        #Name of output
        name = key

        #Output directory
        outputdir = dir_path+'/resultsGPU3/'+name

        #Board Size
        width = value['board'][0] #y axis
        lenght = value['board'][1] #x axis

        PieceQuantity = 1 
        freq = 1 #gx=gy=1

        subJackobs = {}
        for poly in value['set']:
            subJackobs[str(poly)] = inputdataset[str(poly)]
        
        polygons = run_js_script_function('file://'+dir_path.replace('\\','/')+'/scriptpy.js', 'calcNFP_INFP', [subJackobs, lenght,width])

        if not os.path.exists(outputdir):
            os.mkdir(outputdir)
        nfp_infp = json.dumps(polygons, indent=2)
        with open(outputdir+'/nfp-infp '+name+'.json', 'w') as file:
            file.write(nfp_infp)

        board = copy.deepcopy(polygons['rect'])
        del polygons['rect']

        num_polygon = len(polygons)


        total = 0
        total_area = 0
        if PieceQuantity is not None:
            if isinstance(PieceQuantity, dict):
                for key, value in PieceQuantity.items():
                    polygons[key]['QUANTITY'] = PieceQuantity[key]
            else:
                for key, value in polygons.items():
                    polygons[key]['QUANTITY'] = PieceQuantity

        for key, value in polygons.items():
            total += polygons[key]['QUANTITY']



        LayerPoly = []
        LayerOfpoint = []
        valIndex = 1
        layer = 0
        PolyIndex = 0
        PolyIndexDict = {}
        layer_poly_ids = {}
        maxnfp = 0              #Error check
        references = {}
        print("generating each layer of points..")

        AccIndex = 0
        for key, value in polygons.items():
            Nfpsdict = {}

            poly = key

            PolyIndexDict[PolyIndex] = poly

            for nfpoly in polygons[poly]['nfps']:
                Nfpsdict[nfpoly['POLYGON']] = nfpoly
                del Nfpsdict[nfpoly['POLYGON']]['POLYGON']
                if len(nfpoly['VERTICES']) > maxnfp:
                    maxnfp = len(nfpoly['VERTICES'])
            polygons[poly]['nfpdict'] = Nfpsdict
            MainPiece = polygons[poly]
            innerfit = MainPiece["innerfit"]
            Nfps = MainPiece["nfps"]
            quantity = MainPiece.get("QUANTITY",1)
            #print("polygon: ",poly, "Quantity:",quantity)
            for i in range(quantity):
                boardPoints = generatePoints(board,freq,valIndex)
                valIndex += len(boardPoints)
                innerpoint = []
                dictBoardPolyinnerFit = dictpoly(innerfit)
                #Filter for innerfit for polygon
                for point in boardPoints:
                    res = geometryUtil.point_in_polygon(point,dictBoardPolyinnerFit)
                    if res != False:
                        innerpoint.append([point['x'],point['y'],point['id']])
                for i in range(0,len(innerpoint)):
                    innerpoint[i][2] = i+AccIndex

                AccIndex += len(innerpoint)
                #innerpoint = np.array(innerpoint,dtype=np.float32)
                #print("there is a",len(innerpoint)," of innerfit points! ")
                LayerOfpoint.append({"POLYGON":poly,"InnerFitPoints":innerpoint,"Layer":layer})
                LayerPoly.append((layer,poly))
                layer_poly_ids[layer] = PolyIndex
                layer += 1
            #print(polygons[poly]['VERTICES'])
            references[PolyIndex] = (polygons[poly]['VERTICES'][0]['x'],polygons[poly]['VERTICES'][0]['y'])
            PolyIndex += 1

        #Check if max exceeds the the constant NFP
        if maxnfp > MAX_NFP_VERTICES:
            raise Exception(f"SIZE ERROR: nfp size ({maxnfp}) exceeds max size of the MAX_NFP_VERTICES({MAX_NFP_VERTICES})")

    
        print("generating the graph..\n")
        ntXgraphAll = nk.Graph()
        ntXgraphInterLayer = nk.Graph()
        ntXgraphComplete = nk.Graph()
        EdgeArray = []
        EdgeIndex = []

        print("Generating complete graph..")
        for mainLayer in LayerOfpoint:
            #make complete graph
            print("generating graph of Layer: ",mainLayer["Layer"])
            points_values = np.array([point[2] for point in mainLayer["InnerFitPoints"]], dtype=int)
            EdgeIndex.append([points_values.min(),points_values.max()])

            newArr = makeFullGraph(mainLayer["InnerFitPoints"])
            EdgeArray.append(newArr)

        EdgeArray = np.vstack(EdgeArray)
        print("adding results to the graph..")
        ntXgraphAll.addEdges((np.array(EdgeArray[:,0]),np.array(EdgeArray[:,1])), addMissing=True)
        ntXgraphComplete.addEdges((np.array(EdgeArray[:,0]),np.array(EdgeArray[:,1])), addMissing=True)
        print("adding complete!")

        num_groups = len(LayerOfpoint)
        # Create pairs to process (excluding self-pairs)
        all_pairs = [(i,j) for i in range(num_groups) 
                     for j in range(num_groups) if i != j]
        polygonPairs = [(i,j) for i in range(num_polygon)
                     for j in range(num_polygon)]
        print("number of pairs: ",len(all_pairs))
        nfpPair = {}
        nfpPairSize = {}
        for i,j in polygonPairs:
            polyA = PolyIndexDict[i]
            polyB = PolyIndexDict[j]
            listpoint = []
            for v in polygons[polyA]['nfpdict'][polyB]['VERTICES']:
                listpoint.append([v['x'],v['y']])

            nfpPair[(i,j)] = np.array(listpoint,dtype=np.float32)
            nfpPairSize[(i,j)] = nfpPair[(i,j)].shape[0]
        print("Generating NFP-graph..")
        results = process_group_pairs(all_pairs,LayerOfpoint,layer_poly_ids,nfpPair,nfpPairSize,references)
        print("NFP-graph generation complete!")
        print("filtering the duplicates..")
        progress_dict = {}
        total_results = []
        for k in all_pairs:
            if (k[1],k[0]) in progress_dict.keys() or (k[0],k[1]) in progress_dict.keys():
                continue
            result_pair = np.concatenate([results[(k[0],k[1])],results[(k[1],k[0])]])
            del results[(k[0],k[1])]
            del results[(k[1],k[0])]

            if result_pair.shape[0] > 1_000_000:
                total_results.append(remove_symmetric_duplicates_gpu(result_pair))
            else:
                total_results.append(remove_symmetric_duplicates_cpu(result_pair))

            progress_dict[(k[1],k[0])] = 1
            progress_dict[(k[0],k[1])] = 1
            gc.collect()
        del results
        del progress_dict
        total_results = np.concatenate(total_results)
        print("result shape: ",total_results.shape)



        print("adding into graph..")
        ntXgraphAll.addEdges((np.array(total_results[:,0]),np.array(total_results[:,1])), addMissing=True)
        ntXgraphInterLayer.addEdges((np.array(total_results[:,0]),np.array(total_results[:,1])), addMissing=True)
    
        print("widht x lenght: ",width," ",lenght,"nXgraphAll node:",ntXgraphAll.numberOfNodes(),"edges: ",ntXgraphAll.numberOfEdges(),"cliques edges: ",ntXgraphComplete.numberOfEdges(),"NFP-edges: ",ntXgraphInterLayer.numberOfEdges(),"density:",nk.graphtools.density(ntXgraphAll))
        proc_end = time.perf_counter()
        proc_time = proc_end-proc_start


        print("write into file ...")
        update_or_add_with_openpyxl(dir_path+"/results_stat.xlsx",name,proc_time)
        continue
        print("writting into file ...")
        
        

        print("saving into csv")
        final_results = np.concatenate([total_results,EdgeArray])
        start = time.time()
        #pd.DataFrame(final_results).to_csv(outputdir+'/graph '+name+'.csv', index=False, header=False,sep='\t')
        parallel_time = time.time() - start
        print("writting file time: ",parallel_time)
        
        
        
        with open(outputdir+'/pointCoordinate '+name+'.txt', 'w') as file:
            file.write('##format: (Layer,x,y,id)' + '\n')
            for layer in LayerOfpoint:
                for points in layer['InnerFitPoints']:
                    file.write(str(layer['Layer'])+' '+str(points[0])+' '+str(points[1])+' '+str(points[2]) + '\n')

        with open(outputdir+'/LayerPoly'+name+'.txt', 'w') as file:
            file.write('##format:  Layer polygon ' + '\n')
            for layer in LayerPoly:
                file.write(str(layer[0])+'\t'+str(layer[1])+'\n')

        pd.DataFrame(EdgeIndex,columns=['start_index','end_index']).to_csv(outputdir+'/cliques'+name+'.csv', index=False, header=True,sep='\t')
        #start = time.time()
        #with open(outputdir+'/graph '+name+'.csv', 'w', newline='') as csvfile:
        #    spamwriter = csv.writer(csvfile, delimiter='\t',
        #                            quotechar='|', quoting=csv.QUOTE_MINIMAL)
        #    for edge in list(ntXgraphAll.iterEdges()):
        #        spamwriter.writerow([edge[0],edge[1]])
        #sequential_writein = time.time() - start
        #print("sequential write in time: ",sequential_writein)
        #print("nXgraphAll node:",ntXgraphAll.numberOfNodes(),"edges: ",ntXgraphAll.numberOfEdges(),"cliques edges: ",ntXgraphComplete.numberOfEdges(),"NFP-edges: ",ntXgraphInterLayer.numberOfEdges(),"density:",nk.graphtools.density(ntXgraphAll))

        with open(outputdir+'/metadata'+name+'.csv', 'w', newline='') as csvfile:
            spamwriter = csv.writer(csvfile, delimiter='\t',
                                    quoting=csv.QUOTE_MINIMAL)
            spamwriter.writerow(["Name :",str(name)])
            spamwriter.writerow(["Total Pieces :",total])
            spamwriter.writerow(["Type of Pieces :",len(polygons)])
            spamwriter.writerow(["Board width",str(width)])
            spamwriter.writerow(["Board length:",str(lenght)])
            spamwriter.writerow(["Number of Nodes:",ntXgraphAll.numberOfNodes()])
            spamwriter.writerow(["Number of Edges:",ntXgraphAll.numberOfEdges()])
            spamwriter.writerow(["Number of Clique Edges:",ntXgraphComplete.numberOfEdges()])
            spamwriter.writerow(["Intra Layer Edges:",ntXgraphInterLayer.numberOfEdges()])

    
    #import code
    #code.interact(local=locals())
