import json
import os
import geometryUtil
import copy
import networkx as nx
import csv
from pathlib import Path
from NFPLib import *
import time
import openpyxl
from openpyxl import Workbook


def update_or_add_with_openpyxl(filename, name, value):
    """Minimal update or add row using openpyxl with 2 decimal formatting"""
    
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    # Look for existing name in column 1
    found_row = None
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value and str(cell_value).lower() == str(name).lower():
            found_row = row
            break
    
    if found_row:
        # Update existing value in column 2
        target_row = found_row
        print(f"Updated {name}: {value}")
    else:
        # Add new row
        target_row = ws.max_row + 1
        ws.cell(row=target_row, column=1, value=name)
        print(f"Added {name}: {value}")
    
    # Write value
    value_cell = ws.cell(row=target_row, column=2, value=value)
    if isinstance(value, float):
        value_cell.number_format = '0.00'
    
    wb.save(filename)


def polygon_area(polygon):
    """
    Calculate the area of a polygon, assuming no self-intersections.
    A negative area indicates counter-clockwise winding direction.
    
    Args:
        polygon: List of points, where each point is a dict with 'x' and 'y' keys
    
    Returns:
        float: Area of the polygon
    """
    area = 0
    j = len(polygon) - 1  # Start j at the last vertex
    
    for i in range(len(polygon)):
        area += (polygon[j]['x'] + polygon[i]['x']) * (polygon[j]['y'] - polygon[i]['y'])
        j = i  # j follows i
    
    return 0.5 * abs(area)

if __name__ == "__main__":
    
    dir_path = os.path.dirname(os.path.realpath(__file__))

    blaz = "/datasetJson/json/blaz.json"
    shapes = "/datasetJson/json/shapes.json"
    shirts = "/datasetJson/json/shirts.json"
    dagli = "/datasetJson/json/dagli.json"
    fu = "/datasetJson/json/fu.json"
    fu5 = "/datasetJson/json/fu5.json"
    fu6 = "/datasetJson/json/fu6.json"
    fu7 = "/datasetJson/json/fu7.json"
    fu8 = "/datasetJson/json/fu8.json"
    fu9 = "/datasetJson/json/fu9.json"
    fu10 = "/datasetJson/json/fu10.json"
    jakobs1 = "/datasetJson/json/jakobs1.json"
    jakobs2 = "/datasetJson/json/jakobs2.json"
    poly1a = "/datasetJson/json/poly1a.json"
    poly1b = "/datasetJson/json/poly1b.json"
    poly1c = "/datasetJson/json/poly1c.json"
    poly1d = "/datasetJson/json/poly1d.json"
    poly1e = "/datasetJson/json/poly1e.json"
    three = "/datasetJson/json/three.json"
    rco = "/datasetJson/json/rco.json"
    artif = "/datasetJson/json/artif.json"
    #dataset struct:
    #{
    #   "name_polygon":{
    #       "VERTICES":[list_of_vertices],
    #       "QUANTITY":quantity,
    #       "NUMBER OF VERTICES": number
    #
    #}
    ##J

    
    datasetThree = [
        {
            "dataset":three,
            "outputName":"three",
            "quantity":1,
            "width":7,
            "length":7            
        },
        {
            "dataset":three,
            "outputName":"threep2",
            "quantity":2,
            "width":7,
            "length":11
            
        },
        {
            "dataset":three,
            "outputName":"threep2w9",
            "quantity":2,
            "width":9,
            "length":9
            
        },
        {
            "dataset":three,
            "outputName":"threep3",
            "quantity":3,
            "width":7,
            "length":16
        },
        {
            "dataset":three,
            "outputName":"threep3w9",
            "quantity":3,
            "width":9,
            "length":13
        }
    ]

    datasetshapes = [
        {
            "dataset":shapes,
            "outputName":"shapes4-small",
            "quantity":1,
            "width":13,
            "length":24            
        },
        {
            "dataset":shapes,
            "outputName":"shapes8",
            "quantity":2,
            "width":20,
            "length":28            
        },
        {
            "dataset":shapes,
            "outputName":"shapes2",
            "quantity":2,
            "width":40,
            "length":16            
        },
        {
            "dataset":shapes,
            "outputName":"shapes4",
            "quantity":4,
            "width":40,
            "length":28            
        },
        {
            "dataset":shapes,
            "outputName":"shapes5",
            "quantity":5,
            "width":40,
            "length":35            
        },
        {
            "dataset":shapes,
            "outputName":"shapes7",
            "quantity":7,
            "width":40,
            "length":48            
        },
        {
            "dataset":shapes,
            "outputName":"shapes9",
            "quantity":
                {  
                    'PIECE 1':9,
                    'PIECE 2':7,
                    'PIECE 3':9,
                    'PIECE 4':9
                },
            "width":40,
            "length":54            
        },
        {
            "dataset":shapes,
            "outputName":"shapes15",
            "quantity":
                {
                    'PIECE 1':15,
                    'PIECE 2':7,
                    'PIECE 3':9,
                    'PIECE 4':12
                },
            "width":40,
            "length":67            
        },
    ]

    datasetblaz = [
        {
            "dataset":blaz,
            "outputName":"blasz2",
            "quantity":{
                'PIECE 1':5,
                'PIECE 2':5,
                'PIECE 3':5,
                'PIECE 4':5,
                'PIECE 5':0,
                'PIECE 6':0,
                'PIECE 7':0
            },
            "width":15,
            "length":27
        },
        {
            "dataset":blaz,
            "outputName":"BLAZEWICZ1",
            "quantity":1,
            "width":15,
            "length":8            
        },
        {
            "dataset":blaz,
            "outputName":"BLAZEWICZ2",
            "quantity":2,
            "width":15,
            "length":16            
        },
        {
            "dataset":blaz,
            "outputName":"BLAZEWICZ3",
            "quantity":3,
            "width":15,
            "length":22            
        },
        {
            "dataset":blaz,
            "outputName":"BLAZEWICZ4",
            "quantity":4,
            "width":15,
            "length":29           
        },
        {
            "dataset":blaz,
            "outputName":"BLAZEWICZ5",
            "quantity":5,
            "width":15,
            "length":36           
        },
    ]

    datasetRCO = [
        {
            "dataset":rco,
            "outputName":"RCO1",
            "quantity":1,
            "width":15,
            "length":8           
        },
        {
            "dataset":rco,
            "outputName":"RCO2",
            "quantity":2,
            "width":15,
            "length":17           
        },
        {
            "dataset":rco,
            "outputName":"RCO3",
            "quantity":3,
            "width":15,
            "length":25           
        },
        {
            "dataset":rco,
            "outputName":"RCO4",
            "quantity":4,
            "width":15,
            "length":29           
        },
        {
            "dataset":rco,
            "outputName":"RCO5",
            "quantity":5,
            "width":15,
            "length":41           
        },
    ]

    datasetArtif = [
        {
            "dataset":artif,
            "outputName":"artif1_2",
            "quantity":{
                'piece 1':1,
                'piece 2':1,
                'piece 3':1,
                'piece 4':2,
                'piece 5':2,
                'piece 6':2,
                'piece 7':2,
                'piece 8':2
            },
            "width":27,
            "length":8           
        },
        {
            "dataset":artif,
            "outputName":"artif2_4",
            "quantity":{
                'piece 1':2,
                'piece 2':2,
                'piece 3':2,
                'piece 4':4,
                'piece 5':4,
                'piece 6':4,
                'piece 7':4,
                'piece 8':4
            },
            "width":27,
            "length":14           
        },
        {
            "dataset":artif,
            "outputName":"artif3_6",
            "quantity":{
                'piece 1':3,
                'piece 2':3,
                'piece 3':3,
                'piece 4':6,
                'piece 5':6,
                'piece 6':6,
                'piece 7':6,
                'piece 8':6
            },
            "width":27,
            "length":20           
        },
        {
            "dataset":artif,
            "outputName":"artif4_8",
            "quantity":{
                'piece 1':4,
                'piece 2':4,
                'piece 3':4,
                'piece 4':8,
                'piece 5':8,
                'piece 6':8,
                'piece 7':8,
                'piece 8':8
            },
            "width":27,
            "length":28           
        },
        {
            "dataset":artif,
            "outputName":"artif5_10",
            "quantity":{
                'piece 1':5,
                'piece 2':5,
                'piece 3':5,
                'piece 4':10,
                'piece 5':10,
                'piece 6':10,
                'piece 7':10,
                'piece 8':10
            },
            "width":27,
            "length":34
        },
        {
            "dataset":artif,
            "outputName":"artif6_12",
            "quantity":{
                'piece 1':6,
                'piece 2':6,
                'piece 3':6,
                'piece 4':12,
                'piece 5':12,
                'piece 6':12,
                'piece 7':12,
                'piece 8':12
            },
            "width":27,
            "length":41
        },
        {
            "dataset":artif,
            "outputName":"artif7_14",
            "quantity":{
                'piece 1':7,
                'piece 2':7,
                'piece 3':7,
                'piece 4':14,
                'piece 5':14,
                'piece 6':14,
                'piece 7':14,
                'piece 8':14
            },
            "width":27,
            "length":48
        },
        {
            "dataset":artif,
            "outputName":"artif",
            "quantity":{
                'piece 1':8,
                'piece 2':8,
                'piece 3':8,
                'piece 4':15,
                'piece 5':15,
                'piece 6':15,
                'piece 7':15,
                'piece 8':15
            },
            "width":27,
            "length":53
        },
    ]

    datasetshirts = [
        {
            "dataset":shirts ,
            "outputName":"shirts1_2",
            "quantity":{
                'PIECE 1':1,
                'PIECE 2':1,
                'PIECE 3':1,
                'PIECE 4':2,
                'PIECE 5':2,
                'PIECE 6':2,
                'PIECE 7':2,
                'PIECE 8':2,
            },
            "width":40,
            "length":13
        },
        {
            "dataset":shirts ,
            "outputName":"shirts2_4",
            "quantity":{
                'PIECE 1':2,
                'PIECE 2':2,
                'PIECE 3':2,
                'PIECE 4':4,
                'PIECE 5':4,
                'PIECE 6':4,
                'PIECE 7':4,
                'PIECE 8':4,
            },
            "width":40,
            "length":20
        },
        {
            "dataset":shirts ,
            "outputName":"shirts3_6",
            "quantity":{
                'PIECE 1':3,
                'PIECE 2':3,
                'PIECE 3':3,
                'PIECE 4':6,
                'PIECE 5':6,
                'PIECE 6':6,
                'PIECE 7':6,
                'PIECE 8':6,
            },
            "width":40,
            "length":26
        },
        {
            "dataset":shirts ,
            "outputName":"shirts4_8",
            "quantity":{
                'PIECE 1':4,
                'PIECE 2':4,
                'PIECE 3':4,
                'PIECE 4':8,
                'PIECE 5':8,
                'PIECE 6':8,
                'PIECE 7':8,
                'PIECE 8':8,
            },
            "width":40,
            "length":35
        },
        {
            "dataset":shirts ,
            "outputName":"shirts5_10",
            "quantity":{
                'PIECE 1':5,
                'PIECE 2':5,
                'PIECE 3':5,
                'PIECE 4':10,
                'PIECE 5':10,
                'PIECE 6':10,
                'PIECE 7':10,
                'PIECE 8':10,
            },
            "width":40,
            "length":42
        },
    ]

    datasetdagli = [
        {
            "dataset":dagli,
            "outputName":"dagli1",
            "quantity":1,
            "width":60,
            "length":25           
        },
    ]

    datasetfu = [
        {
            "dataset":fu5,
            "outputName":"fu5",
            "quantity":1,
            "width":38,
            "length":18           
        },
        {
            "dataset":fu6,
            "outputName":"fu6",
            "quantity":1,
            "width":38,
            "length":24           
        },
        {
            "dataset":fu7,
            "outputName":"fu7",
            "quantity":1,
            "width":38,
            "length":24           
        },
        {
            "dataset":fu8,
            "outputName":"fu8",
            "quantity":1,
            "width":38,
            "length":24           
        },
        {
            "dataset":fu9,
            "outputName":"fu9",
            "quantity":1,
            "width":38,
            "length":29           
        },
        {
            "dataset":fu10,
            "outputName":"fu10",
            "quantity":1,
            "width":38,
            "length":34           
        },
        {
            "dataset":fu,
            "outputName":"fu",
            "quantity":1,
            "width":38,
            "length":38           
        },
    ]



    PieceQuantityShape9 = {  #Dict if we wants specify quantity of each piece (SHAPE9)
        'PIECE 1':9,
        'PIECE 2':7,
        'PIECE 3':9,
        'PIECE 4':9
    }

    PieceQuantityShape15 = {
        'PIECE 1':15,
        'PIECE 2':7,
        'PIECE 3':9,
        'PIECE 4':12
    }

    PieceBlasz2 = {
        'PIECE 1':5,
        'PIECE 2':5,
        'PIECE 3':5,
        'PIECE 4':5,
        'PIECE 5':0,
        'PIECE 6':0,
        'PIECE 7':0,
    }

    shirts1_2 = {
        'PIECE 1':1,
        'PIECE 2':1,
        'PIECE 3':1,
        'PIECE 4':2,
        'PIECE 5':2,
        'PIECE 6':2,
        'PIECE 7':2,
        'PIECE 8':2,
    }

    shirts2_4 = {
        'PIECE 1':2,
        'PIECE 2':2,
        'PIECE 3':2,
        'PIECE 4':4,
        'PIECE 5':4,
        'PIECE 6':4,
        'PIECE 7':4,
        'PIECE 8':4,
    }

    shirts3_6 = {
        'PIECE 1':3,
        'PIECE 2':3,
        'PIECE 3':3,
        'PIECE 4':6,
        'PIECE 5':6,
        'PIECE 6':6,
        'PIECE 7':6,
        'PIECE 8':6,
    }

    shirts4_8 = {
        'PIECE 1':4,
        'PIECE 2':4,
        'PIECE 3':4,
        'PIECE 4':8,
        'PIECE 5':8,
        'PIECE 6':8,
        'PIECE 7':8,
        'PIECE 8':8,
    }

    shirts5_10 = {
        'PIECE 1':5,
        'PIECE 2':5,
        'PIECE 3':5,
        'PIECE 4':10,
        'PIECE 5':10,
        'PIECE 6':10,
        'PIECE 7':10,
        'PIECE 8':10
    }

    poly_jakobs = [
        {
            "dataset":poly1a,
            "outputName":"poly1a",
            "quantity":1,
            "width":40,
            "length":18           
        },
        {
            "dataset":poly1b,
            "outputName":"poly1b",
            "quantity":1,
            "width":40,
            "length":21           
        },
        {
            "dataset":poly1c,
            "outputName":"poly1c",
            "quantity":1,
            "width":40,
            "length":14           
        },
        {
            "dataset":poly1d,
            "outputName":"poly1d",
            "quantity":1,
            "width":40,
            "length":14           
        },
        {
            "dataset":poly1e,
            "outputName":"poly1e",
            "quantity":1,
            "width":40,
            "length":13           
        },
        {
            "dataset":jakobs1,
            "outputName":"jakobs1",
            "quantity":1,
            "width":40,
            "length":13           
        },
    ]
    #datasetThree
    #datasetshapes
    #datasetblaz
    #datasetRCO
    #datasetshirts
    #datasetdagli
    #datasetfu
    #datasetartif
    datasetshapesBig = [
    {
            "dataset":shapes,
            "outputName":"shapes9",
            "quantity":
                {  
                    'PIECE 1':9,
                    'PIECE 2':7,
                    'PIECE 3':9,
                    'PIECE 4':9
                },
            "width":40,
            "length":54            
        },
        {
            "dataset":shapes,
            "outputName":"shapes15",
            "quantity":
                {
                    'PIECE 1':15,
                    'PIECE 2':7,
                    'PIECE 3':9,
                    'PIECE 4':12
                },
            "width":40,
            "length":67            
        }
    ]

    datasetblaz2 = [
        {
            "dataset":blaz,
            "outputName":"blasz2",
            "quantity":{
                'PIECE 1':5,
                'PIECE 2':5,
                'PIECE 3':5,
                'PIECE 4':5,
                'PIECE 5':0,
                'PIECE 6':0,
                'PIECE 7':0
            },
            "width":15,
            "length":27
        },
     ]
    #datasetTotal = datasetblaz[1:]+datasetRCO+datasetdagli+datasetfu+datasetshirts+datasetArtif
    datasetTotal = datasetblaz2
    #Dataset selected for graph generation
    for set in datasetTotal:
        proc_start = time.perf_counter()
        #Dataset selected for graph generation
        #selelected = shapes
        selelected = set['dataset']
        filepath = dir_path+selelected

        #Name of output
        name = set['outputName']
        #Output directory
        outputdir = dir_path+'/results/'+name

        #Board Size
        #width = 40 #y axis
        #length = 28 #x axis

        width = set['width']
        length = set['length']
        #GPU Settings:

        PieceQuantity = set['quantity'] #None if use quantity from dataset
        #PieceQuantity = {  #Dict if we wants specify quantity of each piece (SHAPE9)
        #    'PIECE 1':9,
        #    'PIECE 2':7,
        #    'PIECE 3':9,
        #    'PIECE 4':9
        #}

        #PieceQuantity = {  #Dict if we wants specify quantity of each piece (SHAPE9)
        #    'PIECE 1':9,
        #    'PIECE 2':7,
        #    'PIECE 3':9,
        #    'PIECE 4':9
        #}



        freq = 1 #gx=gy=1
        with open(filepath) as file:
            inputdataset = json.load(file)
        #

        polygons = run_js_script_function('file://'+dir_path.replace('\\','/')+'/scriptpy.js', 'calcNFP_INFP', [inputdataset, length,width])

        #nfp-infit output struct:
        #{
        #   "name_polygon":{
        #       "VERTICES":[list_of_vertices],
        #       "QUANTITY":quantity,
        #       "NUMBER OF VERTICES": number,
        #       "innerfit":[polygon_innerfit],
        #       "nfps":[{"POLYGON": "name_polygon","VERTICES":[list_of_vertices_nfp]}]}
        #
        #}
        #NOTE: There is also a 'rect' polygon added at the end of the file as the board
        if not os.path.exists(outputdir):
            os.mkdir(outputdir)
        nfp_infp = json.dumps(polygons, indent=2)
        with open(outputdir+'/nfp-infp '+name+'.json', 'w') as file:
            file.write(nfp_infp)



        board = copy.deepcopy(polygons['rect'])
        del polygons['rect']

        print(len(polygons))
        total_polygon = 0
        total_area = 0
        if PieceQuantity is not None:
            if isinstance(PieceQuantity, dict):
                for key, value in PieceQuantity.items():
                    polygons[key]['QUANTITY'] = PieceQuantity[key]
            else:
                for key, value in polygons.items():
                    polygons[key]['QUANTITY'] = PieceQuantity

        for key, value in polygons.items():
            total_area += polygons[key]['QUANTITY']*polygon_area(polygons[key]['VERTICES'])
            total_polygon += polygons[key]['QUANTITY']

        

        LayerPoly = []
        LayerOfpoint = []
        valIndex = 1
        layer = 1
        print("generating each layer of points..")
        for key, value in polygons.items():
            poly = key
            MainPiece = polygons[poly]
            innerfit = MainPiece["innerfit"]
            Nfps = MainPiece["nfps"]
            quantity = MainPiece.get("QUANTITY",1)
            print("polygon: ",key, "Quantity:",quantity)
            for i in range(quantity):
                boardPoints = generatePoints(board,freq,valIndex)
                valIndex += len(boardPoints)
                innerpoint = []
                dictBoardPolyinnerFit = dictpoly(innerfit)
                #Filter for innerfit for polygon
                for point in boardPoints:
                    res = geometryUtil.point_in_polygon(point,dictBoardPolyinnerFit)
                    if res != False:
                        innerpoint.append(point)
                #print("there is a",len(innerpoint)," of innerfit points! ")
                LayerOfpoint.append({"POLYGON":key,"InnerFitPoints":innerpoint,"Layer":layer})
                LayerPoly.append((layer,key))
                layer += 1



        print("generating the graph..")
        #ntXgraphAll = nx.Graph()
        ntXgraphInterLayer = nx.Graph()
        ntXgraphComplete = nx.Graph()
        for mainLayer in LayerOfpoint:
            LayerGraph = {}
            #make full graph
            print("generating graph of Layer: ",mainLayer["Layer"])
            #completeGraph = makeFullGraph(mainLayer["InnerFitPoints"])
            #for g in completeGraph:
            #    ntXgraphAll.add_edge(g[0],g[1])
            #    ntXgraphComplete.add_edge(g[0],g[1])

            #FinalGraph.extend(completeGraph)
            #LayerGraph['completeGraph'] = completeGraph
            LayerGraph["Layer"] = mainLayer["Layer"]

            MainPoly = mainLayer['POLYGON']
            print("iterating point on Layer: ",mainLayer["Layer"])
            resList = MakePointGraph_process_pool(mainLayer["InnerFitPoints"],LayerOfpoint,mainLayer,polygons,dictBoardPolyinnerFit)
            for res in resList[0]:
                for p in res:
                    #ntXgraphAll.add_edge(p['idA'],p['idB'])
                    ntXgraphInterLayer.add_edge(p['idA'],p['idB'])

        proc_end = time.perf_counter()
        proc_time = proc_end-proc_start


        print("write into file ...")
        update_or_add_with_openpyxl(dir_path+"/results_stat.xlsx",name,proc_time)
        #del ntXgraphAll
        del ntXgraphInterLayer
        del ntXgraphComplete
        #with open(outputdir+'/pointCoordinate '+name+'.txt', 'w') as file:
        #    file.write('##format: (Layer,x,y,id)' + '\n')
        #    for layer in LayerOfpoint:
        #        for points in layer['InnerFitPoints']:
        #            file.write(str(layer['Layer'])+' '+str(points['x'])+' '+str(points['y'])+' '+str(points['id']) + '\n')
#
        #with open(outputdir+'/LayerPoly'+name+'.txt', 'w') as file:
        #    file.write('##format:  Layer polygon ' + '\n')
        #    for layer in LayerPoly:
        #        file.write(str(layer[0])+'\t'+str(layer[1])+'\n')
#
#
        #with open(outputdir+'/graph '+name+'.csv', 'w', newline='') as csvfile:
        #    spamwriter = csv.writer(csvfile, delimiter='\t',
        #                            quotechar='|', quoting=csv.QUOTE_MINIMAL)
        #    for edge in list(ntXgraphAll.edges()):
        #        spamwriter.writerow([edge[0],edge[1]])
#
        #print("nXgraphAll node:",ntXgraphAll.number_of_nodes(),"edges: ",ntXgraphAll.number_of_edges(),"cliques edges: ",ntXgraphComplete.number_of_edges(),"NFP-edges: ",ntXgraphInterLayer.number_of_edges(),"density:",nx.density(ntXgraphAll))
#
        #with open(outputdir+'/metadata'+name+'.csv', 'w', newline='') as csvfile:
        #    spamwriter = csv.writer(csvfile, delimiter='\t',
        #                            quoting=csv.QUOTE_MINIMAL)
        #    spamwriter.writerow(["Name :",str(name)])
        #    spamwriter.writerow(["Total Pieces :",total_polygon])
        #    spamwriter.writerow(["Type of Pieces :",len(polygons)])
        #    spamwriter.writerow(["Board lenght x lenght:", str(width)+' '+ str(length)])
        #    spamwriter.writerow(["Number of Nodes:",ntXgraphAll.number_of_nodes()])
        #    spamwriter.writerow(["Number of Edges:",ntXgraphAll.number_of_edges()])
        #    spamwriter.writerow(["Number of Clique Edges:",ntXgraphComplete.number_of_edges()])
        #    spamwriter.writerow(["Intra Layer Edges:",ntXgraphInterLayer.number_of_edges()])
        

        #import code
        #code.interact(local=locals())
